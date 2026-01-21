#
#     ____                  __________
#    / __ \_   _____  _____/ __/ / __ \_      __
#   / / / / | / / _ \/ ___/ /_/ / / / / | /| / /
#  / /_/ /| |/ /  __/ /  / __/ / /_/ /| |/ |/ /
#  \____/ |___/\___/_/  /_/ /_/\____/ |__/|__/
# 
#  The copyright indication and this authorization indication shall be
#  recorded in all copies or in important parts of the Software.
# 
#  @link https://github.com/0verfl0w767
#
import subprocess
import sys

def install_requirements(requirements_file):
    try:
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", "-r", requirements_file],
            text=True,
            capture_output=True,
            check=True,
        )
        print(result.stdout)
    except subprocess.CalledProcessError as e:
        print(e)
    except Exception as e:
        print(e)

install_requirements("requirements.txt")

import os
import json
import openpyxl
from datetime import datetime
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.utils import get_column_letter

from utils.Logger import Logger

LOGGER = Logger()

LOGGER.logo()

RED_TEXT = "\033[1;91m"
GREEN_TEXT = "\033[1;92m"
YELLOW_TEXT = "\033[1;93m"
RED_B_TEXT = "\033[41m"
GREEN_B_TEXT = "\033[42m"
BLUE_B_TEXT = "\033[44m"
PURPLE_B_TEXT = "\033[45m"
RESET_TEXT = "\033[0m"

year = ""
semester = ""
folder_path = ""
data_path = ""
data_info = {
  "grad-count": 0,
  "prev-all-count": 0,
  "pres-all-count": 0,
  "prev-pres-not-found": 0,
  "prev-pres-not-found-list": [],
  "prev-not-found": 0,
  "prev-not-found-list": [],
  "prev-found": 0,
  "prev-found-list": [],
  "pres-warning-1": 0,
  "pres-warning-2": 0,
  "check-warning": 0
}

if not os.path.exists("config.json"):
  configData = {
    "year": "",
    "semester": "",
    "folder-path": "",
    "data-path": "",
  }
  
  with open("config.json", "w", encoding="utf-8") as f:
    json.dump(configData, f, ensure_ascii=False, indent=2)

with open("config.json", "r", encoding="utf-8") as f:
  JSON_DATA = json.load(f)
  year = JSON_DATA["year"]
  semester = JSON_DATA["semester"]
  folder_path = JSON_DATA["folder-path"]
  data_path = JSON_DATA["data-path"]

if not os.path.exists(os.path.join(os.path.dirname(__file__), folder_path)):
  os.makedirs(os.path.join(os.path.dirname(__file__), folder_path))

REAL_PATH = os.path.join(data_path, year, semester)
ABS_PATH_1 = os.path.abspath(REAL_PATH)

FILE_NAME = year + "년 " + semester + " 시간표"
XLSX_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), folder_path + FILE_NAME + ".xlsx"))
JSON_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), folder_path + FILE_NAME + ".json"))

LOGGER.info(FILE_NAME)

allAPI = []

for COLLEGE in os.listdir(ABS_PATH_1):
  ABS_PATH_2 = os.path.abspath(os.path.join(ABS_PATH_1, COLLEGE))
  
  if COLLEGE == "수강편람" or COLLEGE == "전체대학" or COLLEGE == "학부(과).json":
    continue
  
  for UNDERGRADUATE in os.listdir(ABS_PATH_2):
    ABS_PATH_3 = os.path.abspath(os.path.join(ABS_PATH_2, UNDERGRADUATE))
    GRAD_NAME = os.path.splitext(UNDERGRADUATE)[0]
    
    LOGGER.info("┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
    LOGGER.info("")
    LOGGER.info(" > 단과대학: " + COLLEGE)
    LOGGER.info(" > 학부(과): " + GRAD_NAME)
    
    with open(ABS_PATH_3, "r", encoding="utf-8") as f:
      PRES_DATA = json.load(f)
      PREV_YEAR = int(ABS_PATH_3.split("\\")[7]) - 1
      PREV_PATH = ABS_PATH_3.replace(year, str(PREV_YEAR))
      
      try:
        with open(PREV_PATH, "r", encoding="utf-8") as f:
          PREV_DATA = json.load(f)
          PRES_COUNT = len(PRES_DATA["api"])
          PREV_COUNT = len(PREV_DATA["api"])
          
          data_info["grad-count"] += 1
          
          if not PRES_DATA["api"]:
            if not PREV_DATA["api"]:
              data_info["prev-pres-not-found"] += 1
              data_info["prev-pres-not-found-list"].append(GRAD_NAME)
              LOGGER.info(" > 강의 계획서가 " + RED_TEXT + "확인되지 않음. (이전 연도 확인되지 않음.)")
            else:
              data_info["prev-found"] += 1
              data_info["prev-found-list"].append(GRAD_NAME)
              LOGGER.info(" > 강의 계획서가 " + RED_TEXT + "확인되지 않음. " + GREEN_TEXT + "(이전 연도 " + str(PREV_COUNT) + "개 확인됨.)")
            # continue
          else:
            MSG = "(" + str(PRES_COUNT - PREV_COUNT) + "개 증가)" if PRES_COUNT > PREV_COUNT else "(변화 없음)" if PRES_COUNT == PREV_COUNT else "(" + str(PREV_COUNT - PRES_COUNT) + "개 감소)"
            
            if not PREV_DATA["api"]:
              data_info["prev-not-found"] += 1
              data_info["prev-not-found-list"].append(GRAD_NAME)
              LOGGER.info(" > 강의 계획서가 " + str(PRES_COUNT) + "개 " + GREEN_TEXT + "확인됨. " + RED_TEXT + "(이전 연도 확인되지 않음.) " + MSG)
            else:
              LOGGER.info(" > 강의 계획서가 " + str(PRES_COUNT) + "개 " + GREEN_TEXT + "확인됨. " + YELLOW_TEXT + "(이전 연도 " + str(PREV_COUNT) + "개 확인됨.) " + MSG)
      except:
        PRES_COUNT = len(PRES_DATA["api"])
      
      data_info["pres-all-count"] += PRES_COUNT
      data_info["prev-all-count"] += PREV_COUNT
      
      MANUAL_COUNT = 0
      
      dfoundList = []
      mfoundList = []
      
      direction = {}
      manual = {}
      
      for realData in PRES_DATA["api"]:
        if (not realData["교수명"] or realData["교수명"] == "미지정") and (not realData["수업시간"] or realData["수업시간"] == "미지정"):
          data_info["pres-warning-1"] += 1
        
        if not realData["장소"] or realData["장소"] == "미지정":
          data_info["pres-warning-2"] += 1
        
        realData["비고"] = realData["이수구분"] if realData["이수구분"] == "교직필수" else realData["영역구분"]
        realData["팀티칭여부"] = ""
        realData["단과대학"] = COLLEGE
        
        dfoundList.append(realData["강좌번호"])
        
        if realData["강좌번호"] in direction:
          direction[realData["강좌번호"]]["count"] += 1
        else:
          direction[realData["강좌번호"]] = {}
          direction[realData["강좌번호"]]["과목명"] = realData["과목명"]
          direction[realData["강좌번호"]]["count"] = 1
        
        manual = {}
        
        MANUAL_PATH = os.path.abspath(os.path.join(ABS_PATH_1, "수강편람", UNDERGRADUATE))
        
        with open(MANUAL_PATH, "r", encoding="utf-8") as f:
          MANUAL_DATA = json.load(f)
          MANUAL_COUNT = len(MANUAL_DATA["api"])
          
          for newData in MANUAL_DATA["api"]:
            if newData["강좌번호"] == realData["강좌번호"]:
              realData["비고"] = newData["비고"] if realData["비고"] == "" else realData["비고"] if newData["비고"] == "" else realData["비고"] + ", " + newData["비고"]
              # realData["비고"] = realData["비고"] + ", " + newData["비고"] if realData["비고"] != "" and newData["비고"] != "" else newData["비고"]
              realData["팀티칭여부"] = newData["팀티칭여부"]
            
            if newData["강좌번호"] in manual:
              manual[newData["강좌번호"]]["count"] += 1
            else:
              manual[newData["강좌번호"]] = {}
              manual[newData["강좌번호"]]["과목명"] = newData["과목명"]
              manual[newData["강좌번호"]]["count"] = 1

        del realData['순번']
        del realData['수업시간/장소']
        allAPI.append(realData)
      
      directionCount = 0
      
      for key, value in direction.items():
        if value["count"] > 1:
          directionCount += value["count"] - 1
          LOGGER.warning(" >> " + RED_TEXT + "강의 계획서에서 중복된 강좌가 " + str(value["count"]) + "개 발견되었습니다. 강좌번호: " + key + ", 과목명: " + value["과목명"])
      
      if MANUAL_COUNT != 0:
        with open(MANUAL_PATH, "r", encoding="utf-8") as f:
          MANUAL_DATA = json.load(f)
          
          for newData in MANUAL_DATA["api"]:
            mfoundList.append(newData["강좌번호"])
            
            if not newData["강좌번호"] in dfoundList:
              newData["단과대학"] = COLLEGE
              del newData['순번']
              allAPI.append(newData)
              LOGGER.warning(" >> " + RED_TEXT + "강의 계획서에서 누락된 강좌가 발견되었습니다. 강좌번호: " + newData["강좌번호"] + ", 과목명: " + newData["과목명"])
              LOGGER.warning(" >> " + PURPLE_B_TEXT + "누락된 강좌가 발견되어 데이터가 추가되었습니다. 강좌번호: " + newData["강좌번호"] + ", 과목명: " + newData["과목명"])
        
        for realData in PRES_DATA["api"]:
          if not realData["강좌번호"] in mfoundList:
            LOGGER.warning(" >> " + RED_TEXT + "수강 편람에서 누락된 강좌가 발견되었습니다. 강좌번호: " + realData["강좌번호"] + ", 과목명: " + realData["과목명"])
      
      MSG = RED_B_TEXT + "(실패) 데이터 값 없음" if PRES_COUNT == 0 or MANUAL_COUNT == 0 else BLUE_B_TEXT + "(통과)" if PRES_COUNT == MANUAL_COUNT else RED_B_TEXT + "(실패) 누락 확인 바람"
      LOGGER.info(" > 수강 편람이 " + str(MANUAL_COUNT) + "개 확인됨. ")
      
      manualCount = 0
      
      for key, value in manual.items():
        if value["count"] > 1:
          manualCount += value["count"] - 1
          LOGGER.warning(" >> " + RED_TEXT + "수강 편람에서 중복된 강좌가 " + str(value["count"]) + "개 발견되었습니다. 강좌번호: " + key + ", 과목명: " + value["과목명"])
      
      LOGGER.info(" > 상태: " + MSG)
      DIRECT_COUNT = PRES_COUNT - directionCount
      MANUAL_COUNT = MANUAL_COUNT - manualCount
      D_M = DIRECT_COUNT - MANUAL_COUNT
      
      MSG = ""
      
      if D_M == 0:
        MSG = BLUE_B_TEXT + "(통과)"
      else:
        MSG = RED_B_TEXT + "(실패)"
        data_info["check-warning"] += 1
      
      LOGGER.info(" > 진단: " + MSG + ", " + str(DIRECT_COUNT) + " - " + str(MANUAL_COUNT) + " = " + str(D_M))
    
    LOGGER.info("")
    LOGGER.info("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")

now = datetime.now()
days = ["월", "화", "수", "목", "금", "토", "일"]

apiJson = {}
apiJson["year"] = year
apiJson["semester"] = semester
apiJson["time"] = now.strftime(f"'%y.%m.%d.({days[now.weekday()]})")
apiJson["api"] = sorted(allAPI, key=lambda i: (i["단과대학"], i["학부(과)"], 1 if i["과목명"] != "채플" else -1, int(i["학년"]), int(i["학점"]), i["과목명"], int(i["강좌번호"])))

with open(JSON_PATH, "w", encoding="utf-8") as f:
  json.dump(apiJson, f, ensure_ascii=False, indent=2)

excelWB = openpyxl.Workbook()
sheet = excelWB.active

sheet.column_dimensions["A"].width = 8 # 강좌번호
sheet.column_dimensions["B"].width = 8 # 과목코드
sheet.column_dimensions["C"].width = 35 # 과목명
sheet.column_dimensions["D"].width = 18 # 학부(과)
sheet.column_dimensions["E"].width = 4 # 학년
sheet.column_dimensions["F"].width = 9 # 이수구분
sheet.column_dimensions["G"].width = 15 # 영역구분
sheet.column_dimensions["H"].width = 4 # 학점
sheet.column_dimensions["I"].width = 10 # 교수명
sheet.column_dimensions["J"].width = 12 # 수업시간
sheet.column_dimensions["K"].width = 30 # 장소
sheet.column_dimensions["L"].width = 12 # 단과대학
sheet.column_dimensions["M"].width = 25 # 비고
sheet.column_dimensions["N"].width = 10 # 팀티칭여부


sheet["A1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["B1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["C1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["D1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["E1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["F1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["G1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["H1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["I1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["J1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["K1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["L1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["M1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["N1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))

sheet.freeze_panes = "A2"

sheet["A1"] = "강좌번호"
sheet["B1"] = "과목코드"
sheet["C1"] = "과목명"
sheet["D1"] = "학부(과)"
sheet["E1"] = "학년"
sheet["F1"] = "이수구분"
sheet["G1"] = "영역구분"
sheet["H1"] = "학점"
sheet["I1"] = "교수명"
sheet["J1"] = "수업시간"
sheet["K1"] = "장소"
sheet["L1"] = "단과대학"
sheet["M1"] = "비고"
sheet["N1"] = "팀티칭여부"

for col in range(1, 15):
  cell = sheet.cell(row=1, column=col)
  cell.font = Font(name="맑은 고딕", bold=True)

with open(JSON_PATH, "r", encoding="utf-8") as f:
  PRES_DATA = json.load(f)
  rowCount = 1
  
  for realData in PRES_DATA["api"]:
    rowCount += 1
    
    sheet["A" + str(rowCount)] = realData["강좌번호"]
    sheet["B" + str(rowCount)] = realData["과목코드"]
    sheet["C" + str(rowCount)] = realData["과목명"]
    sheet["D" + str(rowCount)] = realData["학부(과)"]
    sheet["E" + str(rowCount)] = realData["학년"]
    sheet["F" + str(rowCount)] = realData["이수구분"]
    sheet["G" + str(rowCount)] = realData["영역구분"]
    sheet["H" + str(rowCount)] = realData["학점"]
    sheet["I" + str(rowCount)] = realData["교수명"]
    sheet["J" + str(rowCount)] = realData["수업시간"]
    sheet["K" + str(rowCount)] = realData["장소"]
    sheet["L" + str(rowCount)] = realData["단과대학"]
    sheet["M" + str(rowCount)] = realData["비고"]
    sheet["N" + str(rowCount)] = realData["팀티칭여부"]


# sheet.protection.sheet = True
# sheet.protection.enable()
    
excelWB.save(XLSX_PATH)
excelWB.close()

LOGGER.info("┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
LOGGER.info("")
LOGGER.info(" > JSON PATH: " + JSON_PATH)
LOGGER.info(" > XLSX PATH: " + XLSX_PATH)
LOGGER.info(" > 전체 학부(과): " + str(data_info["grad-count"]) + "개 확인됨.")
LOGGER.info(" > 올해 전체 강의: " + str(data_info["pres-all-count"]) + "개 확인됨.")
LOGGER.info(" > 작년 전체 강의: " + str(data_info["prev-all-count"]) + "개 확인됨.")
LOGGER.info(" > " + str(data_info["prev-found"]) + "개의 학부(과)가 강의 계획서가 없는 것으로 추정됨.")
LOGGER.info(" > " + PURPLE_B_TEXT + ", ".join(data_info["prev-found-list"]))
LOGGER.info(" > " + str(data_info["prev-pres-not-found"]) + "개의 학부(과)가 폐지된 과로 추정됨.")
LOGGER.info(" > " + RED_B_TEXT + ", ".join(data_info["prev-pres-not-found-list"]))
LOGGER.info(" > " + str(data_info["prev-not-found"]) + "개의 학부(과)가 신설된 과로 추정됨.")
LOGGER.info(" > " + GREEN_B_TEXT + ", ".join(data_info["prev-not-found-list"]))

VALUE = (data_info["pres-warning-1"] / data_info["pres-all-count"]) * 100
MSG = BLUE_B_TEXT + "에브리타임 시간표 업데이트 통과" if VALUE < 15 else RED_B_TEXT + "에브리타임 시간표 업데이트 실패"
LOGGER.info(" > " + str(data_info["pres-all-count"]) + "개의 강의 중 교수와 시간이 정해지지 않은 " + str(data_info["pres-warning-1"]) + "개의 강의가 확인됨.")
LOGGER.info(" > " + str(data_info["pres-all-count"]) + "개의 강의 중 장소가 정해지지 않은 " + str(data_info["pres-warning-2"]) + "개의 강의가 확인됨.")
LOGGER.info(" > 상태: " + MSG + " (" + str(int(VALUE)) + "%)")
LOGGER.info(" > 진단: " + str(data_info["check-warning"]) + "개의 학부(과)가 진단에 실패했습니다.")
LOGGER.info("")
LOGGER.info("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
