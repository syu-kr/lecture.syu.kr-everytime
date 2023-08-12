#
#     ____                  __________
#    / __ \_   _____  _____/ __/ / __ \_      __
#   / / / / | / / _ \/ ___/ /_/ / / / / | /| / /
#  / /_/ /| |/ /  __/ /  / __/ / /_/ /| |/ |/ /
#  \____/ |___/\___/_/  /_/ /_/\____/ |__/|__/
# 
#  The copyright indication and this authorization indication shall be
#  recorded in all copies or in important parts of the Software.
# 
#  @author 0verfl0w767
#  @link https://github.com/0verfl0w767
#  @license MIT LICENSE
#
import os
import json
import openpyxl
from openpyxl.styles import Color, PatternFill

from utils.Logger import Logger

LOGGER = Logger()

LOGGER.logo()

RED_TEXT = "\033[1;91m"
GREEN_TEXT = "\033[1;92m"
YELLOW_TEXT = "\033[1;93m"
RED_B_TEXT = "\033[41m"
GREEN_B_TEXT = "\033[42m"
BLUE_B_TEXT = "\033[44m"
PURPLE_B_TEXT = "\033[45m"
RESET_TEXT = "\033[0m"

year = ""
semester = ""
folder_path = ""
data_path = ""
data_info = {
  "grad-count": 0,
  "prev-all-count": 0,
  "pres-all-count": 0,
  "prev-pres-not-found": 0,
  "prev-pres-not-found-list": [],
  "prev-not-found": 0,
  "prev-not-found-list": [],
  "prev-found": 0,
  "prev-found-list": [],
  "pres-warning": 0,
  "check-warning": 0
}

if not os.path.exists("config.json"):
  configData = {
    "year": "",
    "semester": "",
    "folder-path": "",
    "data-path": "",
  }
  
  with open("config.json", "w", encoding = "utf-8") as f:
    json.dump(configData, f, ensure_ascii = False, indent = 2)

with open("config.json", "r", encoding = "utf-8") as f:
  JSON_DATA = json.load(f)
  year = JSON_DATA["year"]
  semester = JSON_DATA["semester"]
  folder_path = JSON_DATA["folder-path"]
  data_path = JSON_DATA["data-path"]

if not os.path.exists(os.path.join(os.path.dirname(__file__), folder_path)):
  os.makedirs(os.path.join(os.path.dirname(__file__), folder_path))

REAL_PATH = data_path + "\\" + year + "\\" + semester
ABS_PATH_1 = os.path.abspath(REAL_PATH)

FILE_NAME = year + "년 " + semester + " 시간표"
XLSX_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), folder_path + FILE_NAME + ".xlsx"))
JSON_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), folder_path + FILE_NAME + ".json"))

LOGGER.info(FILE_NAME)

allAPI = []

for COLLEGE in os.listdir(ABS_PATH_1):
  ABS_PATH_2 = os.path.abspath(ABS_PATH_1 + "\\" + COLLEGE)
  
  if COLLEGE == "수강편람" or COLLEGE == "전체대학" or COLLEGE == "학부(과).json":
    continue
  
  for UNDERGRADUATE in os.listdir(ABS_PATH_2):
    ABS_PATH_3 = os.path.abspath(ABS_PATH_1 + "\\" + "수강편람" + "\\" + UNDERGRADUATE)
    GRAD_NAME = os.path.splitext(UNDERGRADUATE)[0]
    
    LOGGER.info("┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
    LOGGER.info("")
    LOGGER.info(" > 단과대학: " + COLLEGE)
    LOGGER.info(" > 학부(과): " + GRAD_NAME)
    
    with open(ABS_PATH_3, "r", encoding = "utf-8") as f:
      PRES_DATA = json.load(f)
      PREV_YEAR = int(ABS_PATH_3.split("\\")[7]) - 1
      PREV_PATH = ABS_PATH_3.replace(year, str(PREV_YEAR))
      
      MANUAL_COUNT = 0
      
      for realData in PRES_DATA["api"]:
        realData["단과대학"] = COLLEGE
        
        del realData['순번']
        allAPI.append(realData)
    
    LOGGER.info("")
    LOGGER.info("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")

apiJson = {}
apiJson["year"] = year
apiJson["semester"] = semester
apiJson["api"] = sorted(allAPI, key=lambda i: (i["단과대학"], i["학부(과)"], int(i["학년"]), int(i["학점"]), i["과목명"], i["강좌번호"]))

with open(JSON_PATH, "w", encoding = "utf-8") as f:
  json.dump(apiJson, f, ensure_ascii = False, indent = 2)

excelWB = openpyxl.Workbook()
sheet = excelWB.active

sheet.column_dimensions["A"].width = 9 # 강좌번호
sheet.column_dimensions["B"].width = 9 # 과목코드
sheet.column_dimensions["C"].width = 35 # 과목명
sheet.column_dimensions["D"].width = 18 # 학부(과)
sheet.column_dimensions["E"].width = 5 # 학년
sheet.column_dimensions["F"].width = 9 # 이수구분
sheet.column_dimensions["G"].width = 15 # 영역구분
sheet.column_dimensions["H"].width = 5 # 학점
sheet.column_dimensions["I"].width = 7 # 교수명
sheet.column_dimensions["J"].width = 15 # 수업시간
sheet.column_dimensions["K"].width = 30 # 장소
sheet.column_dimensions["L"].width = 12 # 단과대학
# sheet.column_dimensions["M"].width = 25 # 비고

sheet["A1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["B1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["C1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["D1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["E1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["F1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["G1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["H1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["I1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["J1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["K1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["L1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["M1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))
sheet["N1"].fill = PatternFill(fill_type="solid", fgColor=Color("29CDFF"))

sheet.freeze_panes = "A2"

sheet["A1"] = "강좌번호"
sheet["B1"] = "과목코드"
sheet["C1"] = "과목명"
sheet["D1"] = "학부(과)"
sheet["E1"] = "학년"
sheet["F1"] = "이수구분"
sheet["G1"] = "영역구분"
sheet["H1"] = "학점"
sheet["I1"] = "교수명"
sheet["J1"] = "수업시간"
sheet["K1"] = "장소"
sheet["L1"] = "단과대학"
sheet["M1"] = "비고"
sheet["N1"] = "팀티칭여부"

with open(JSON_PATH, "r", encoding = "utf-8") as f:
  PRES_DATA = json.load(f)
  rowCount = 1
  
  for realData in PRES_DATA["api"]:
    rowCount += 1
    
    sheet["A" + str(rowCount)] = realData["강좌번호"]
    sheet["B" + str(rowCount)] = realData["과목코드"]
    sheet["C" + str(rowCount)] = realData["과목명"]
    sheet["D" + str(rowCount)] = realData["학부(과)"]
    sheet["E" + str(rowCount)] = realData["학년"]
    sheet["F" + str(rowCount)] = realData["이수구분"]
    sheet["G" + str(rowCount)] = realData["영역구분"]
    sheet["H" + str(rowCount)] = realData["학점"]
    sheet["I" + str(rowCount)] = realData["교수명"]
    sheet["J" + str(rowCount)] = realData["수업시간"]
    sheet["K" + str(rowCount)] = realData["장소"]
    sheet["L" + str(rowCount)] = realData["단과대학"]
    sheet["L" + str(rowCount)] = realData["단과대학"]
    sheet["M" + str(rowCount)] = realData["비고"]
    sheet["N" + str(rowCount)] = realData["팀티칭여부"]

excelWB.save(XLSX_PATH)
excelWB.close()

LOGGER.info("┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
LOGGER.info("")
LOGGER.info(" > JSON PATH: " + JSON_PATH)
LOGGER.info(" > XLSX PATH: " + XLSX_PATH)
LOGGER.info(" > 전체 학부(과): " + str(data_info["grad-count"]) + "개 확인됨.")
LOGGER.info(" > 올해 전체 강의: " + str(data_info["pres-all-count"]) + "개 확인됨.")
LOGGER.info(" > 작년 전체 강의: " + str(data_info["prev-all-count"]) + "개 확인됨.")
LOGGER.info(" > " + str(data_info["prev-found"]) + "개의 학부(과)가 강의 계획서가 없는 것으로 추정됨.")
LOGGER.info(" > " + PURPLE_B_TEXT + ", ".join(data_info["prev-found-list"]))
LOGGER.info(" > " + str(data_info["prev-pres-not-found"]) + "개의 학부(과)가 폐지된 과로 추정됨.")
LOGGER.info(" > " + RED_B_TEXT + ", ".join(data_info["prev-pres-not-found-list"]))
LOGGER.info(" > " + str(data_info["prev-not-found"]) + "개의 학부(과)가 신설된 과로 추정됨.")
LOGGER.info(" > " + GREEN_B_TEXT + ", ".join(data_info["prev-not-found-list"]))

# VALUE = (data_info["pres-warning"] / data_info["pres-all-count"]) * 100
# MSG = BLUE_B_TEXT + "에브리타임 시간표 업데이트 통과" if VALUE < 15 else RED_B_TEXT + "에브리타임 시간표 업데이트 실패"
# LOGGER.info(" > " + str(data_info["pres-all-count"]) + "개의 강의 중 교수와 시간이 정해지지 않은 " + str(data_info["pres-warning"]) + "개의 강의가 확인됨.")
# LOGGER.info(" > 상태: " + MSG + " (" + str(int(VALUE)) + "%)")
# LOGGER.info(" > 진단: " + str(data_info["check-warning"]) + "개의 학부(과)가 진단에 실패했습니다.")
LOGGER.info("")
LOGGER.info("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")